"""Build the per-video location-breakdown Excel workbook.

Two sheets:
  • "Location Breakdown" — the table the customer asked for
    (Location | Logo | Human % | AI % | Human-AI %), with a metadata header and a
    totals row.
  • "AI % Detail" — the parameters behind every AI %: which criteria were enabled,
    and per location the factor means (size / position / clarity / OBB), the
    recomputed frame weight, segment count, on-screen duration and quality
    exposure that the share is normalised from.

Kept dependency-light: just openpyxl (added to pyproject).
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.pipeline.location_breakdown import AI_CRITERIA

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13)
_TOTAL_FONT = Font(bold=True)
_PCT = "0.00"
_NUM = "0.000"


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_location_workbook(
    *,
    analysis,
    rows: list[dict],
    zone_detail: dict[str, dict],
    enabled: list[str],
) -> bytes:
    label_by_key = {c["key"]: c["label"] for c in AI_CRITERIA}
    enabled_labels = ", ".join(label_by_key.get(k, k) for k in enabled) or "none"

    wb = Workbook()

    # ── Sheet 1: Location Breakdown ───────────────────────────────────────
    ws = wb.active
    ws.title = "Location Breakdown"

    ws["A1"] = "Location Breakdown"
    ws["A1"].font = _TITLE_FONT
    meta = [
        ("Event", analysis.event_name or ""),
        ("Video", analysis.video_name or ""),
        ("Analysed at", analysis.result_json.get("analyzedAt", "")),
        ("AI criteria enabled", enabled_labels),
    ]
    r = 2
    for k, v in meta:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1

    head_row = r + 1
    headers = ["Location", "Logo", "Human %", "AI %", "Human AI %", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=head_row, column=c, value=h)
    _style_header(ws, head_row, len(headers))

    rr = head_row + 1
    h_total = ai_total = 0.0
    for row in rows:
        ws.cell(row=rr, column=1, value=row["locationName"])
        ws.cell(row=rr, column=2, value=row["logo"] or "")
        ws.cell(row=rr, column=3, value=row["humanPercentage"]).number_format = _PCT
        ws.cell(row=rr, column=4, value=row["aiPercentage"]).number_format = _PCT
        hai = row["humanAiPercentage"]
        cell = ws.cell(row=rr, column=5, value=hai)
        if hai is not None:
            cell.number_format = _PCT
        ws.cell(row=rr, column=6, value=row["notes"] or "")
        h_total += row["humanPercentage"] or 0.0
        ai_total += row["aiPercentage"] or 0.0
        rr += 1

    ws.cell(row=rr, column=1, value="Total").font = _TOTAL_FONT
    t3 = ws.cell(row=rr, column=3, value=round(h_total, 2)); t3.font = _TOTAL_FONT; t3.number_format = _PCT
    t4 = ws.cell(row=rr, column=4, value=round(ai_total, 2)); t4.font = _TOTAL_FONT; t4.number_format = _PCT

    _autosize(ws, [22, 20, 12, 12, 14, 28])

    # ── Sheet 2: AI % Detail ──────────────────────────────────────────────
    ws2 = wb.create_sheet("AI % Detail")
    ws2["A1"] = "AI Percentage — parameters"
    ws2["A1"].font = _TITLE_FONT
    ws2["A2"] = "Enabled criteria"
    ws2["A2"].font = Font(bold=True)
    ws2["B2"] = enabled_labels
    ws2["A3"] = (
        "AI % = each anchor zone's share of total quality exposure under the enabled "
        "criteria. Locations sharing an anchor split that share evenly."
    )
    ws2["A3"].font = Font(italic=True, color="6B7280")

    head2 = 5
    headers2 = [
        "Location", "Anchor zone", "Logo", "AI %", "Quality exposure",
        "Detections", "Segments", "On-screen (s)",
        "Mean Size", "Mean Position", "Mean Clarity", "Mean OBB", "Mean frame weight",
    ]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=head2, column=c, value=h)
    _style_header(ws2, head2, len(headers2))

    rr = head2 + 1
    for row in rows:
        d = zone_detail.get(row["anchorId"], {})
        ws2.cell(row=rr, column=1, value=row["locationName"])
        ws2.cell(row=rr, column=2, value=row["anchorId"])
        ws2.cell(row=rr, column=3, value=row["logo"] or "")
        ws2.cell(row=rr, column=4, value=row["aiPercentage"]).number_format = _PCT
        ws2.cell(row=rr, column=5, value=round(d.get("quality", 0.0), 4)).number_format = _NUM
        ws2.cell(row=rr, column=6, value=d.get("detections", 0))
        ws2.cell(row=rr, column=7, value=d.get("segments", 0))
        ws2.cell(row=rr, column=8, value=round(d.get("totalDuration", 0.0), 2)).number_format = _PCT
        ws2.cell(row=rr, column=9, value=round(d.get("meanSize", 0.0), 4)).number_format = _NUM
        ws2.cell(row=rr, column=10, value=round(d.get("meanPos", 0.0), 4)).number_format = _NUM
        ws2.cell(row=rr, column=11, value=round(d.get("meanClarity", 0.0), 4)).number_format = _NUM
        ws2.cell(row=rr, column=12, value=round(d.get("meanObb", 0.0), 4)).number_format = _NUM
        ws2.cell(row=rr, column=13, value=round(d.get("meanFrameWeight", 0.0), 4)).number_format = _NUM
        rr += 1

    _autosize(ws2, [22, 16, 18, 10, 16, 12, 10, 14, 11, 13, 12, 11, 16])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
